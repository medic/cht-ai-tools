#!/usr/bin/env python3
"""
Add a calculated field to an XLSForm file.
Uses openpyxl to modify Excel files.

Usage: python add-xlsform-calculation.py <xlsform.xlsx> <field_name> <calculation> [--before=group_name]

Examples:
    python add-xlsform-calculation.py forms/app/delivery.xlsx needs_urgent_pnc "if(\${risk}='high','yes','no')"
    python add-xlsform-calculation.py forms/app/delivery.xlsx needs_followup "if(\${referred}='yes','yes','no')" --before=group_summary
"""

import sys
import argparse
from pathlib import Path
from copy import copy

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def find_column_index(sheet, column_name):
    """Find the index of a column by name in the header row."""
    for idx, cell in enumerate(sheet[1], start=1):
        if cell.value and str(cell.value).strip().lower() == column_name.lower():
            return idx
    return None


def find_insert_position(sheet, type_col, name_col, before_group=None):
    """
    Find the row number where to insert the new calculation.

    Strategy:
    1. If before_group specified, insert before that group
    2. Otherwise, insert before 'group_summary' or 'summary' group
    3. Otherwise, insert at the end (before any 'end group' that closes the form)
    4. Fallback: append at the very end
    """
    last_content_row = 1
    summary_row = None
    target_group_row = None

    for row_num in range(2, sheet.max_row + 1):
        type_val = sheet.cell(row=row_num, column=type_col).value
        name_val = sheet.cell(row=row_num, column=name_col).value

        if type_val:
            type_str = str(type_val).strip().lower()
            name_str = str(name_val).strip().lower() if name_val else ""

            # Track last row with content
            last_content_row = row_num

            # Check for target group (if specified)
            if before_group and type_str.startswith("begin") and name_str == before_group.lower():
                target_group_row = row_num

            # Check for summary group
            if type_str.startswith("begin") and ("summary" in name_str or name_str == "group_summary"):
                summary_row = row_num

    # Priority: specified group > summary group > end of content
    if target_group_row:
        return target_group_row
    if summary_row:
        return summary_row

    return last_content_row + 1


def add_calculation(filepath, field_name, calculation, before_group=None):
    """Add a calculate row to the XLSForm survey sheet."""

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return {"success": False, "error": f"Cannot open file: {e}"}

    if "survey" not in wb.sheetnames:
        return {"success": False, "error": "No 'survey' sheet found in workbook"}

    sheet = wb["survey"]

    # Find required columns
    type_col = find_column_index(sheet, "type")
    name_col = find_column_index(sheet, "name")
    calc_col = find_column_index(sheet, "calculation")
    label_col = find_column_index(sheet, "label")

    if not type_col:
        return {"success": False, "error": "Could not find 'type' column in survey sheet"}
    if not name_col:
        return {"success": False, "error": "Could not find 'name' column in survey sheet"}

    # Check if calculation column exists, if not create it
    if not calc_col:
        # Find the last column and add calculation column
        calc_col = sheet.max_column + 1
        sheet.cell(row=1, column=calc_col, value="calculation")

    # Check if field already exists
    for row_num in range(2, sheet.max_row + 1):
        name_val = sheet.cell(row=row_num, column=name_col).value
        if name_val and str(name_val).strip() == field_name:
            return {"success": False, "error": f"Field '{field_name}' already exists in the form"}

    # Find insert position
    insert_row = find_insert_position(sheet, type_col, name_col, before_group)

    # Insert a new row
    sheet.insert_rows(insert_row)

    # Copy formatting from the row above (if it exists and has content)
    if insert_row > 2:
        for col in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(row=insert_row + 1, column=col)
            target_cell = sheet.cell(row=insert_row, column=col)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.alignment = copy(source_cell.alignment)

    # Set the values for the new row
    sheet.cell(row=insert_row, column=type_col, value="calculate")
    sheet.cell(row=insert_row, column=name_col, value=field_name)
    sheet.cell(row=insert_row, column=calc_col, value=calculation)

    # Optionally set an empty label (some forms require it)
    if label_col:
        sheet.cell(row=insert_row, column=label_col, value="")

    # Save the workbook
    try:
        wb.save(filepath)
    except Exception as e:
        return {"success": False, "error": f"Cannot save file: {e}"}
    finally:
        wb.close()

    return {
        "success": True,
        "field_name": field_name,
        "calculation": calculation,
        "inserted_at_row": insert_row,
        "file": str(filepath)
    }


def main():
    parser = argparse.ArgumentParser(
        description="Add a calculated field to an XLSForm file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s forms/app/delivery.xlsx needs_urgent_pnc "if(\\${risk}='high','yes','no')"
  %(prog)s delivery.xlsx needs_followup "if(\\${referred}='yes','yes','no')" --before=group_summary

Note: Escape $ signs in the shell with \\$ or use single quotes around the calculation.
        """
    )
    parser.add_argument("xlsform", help="Path to the XLSForm (.xlsx) file")
    parser.add_argument("field_name", help="Name for the new calculated field")
    parser.add_argument("calculation", help="XLSForm calculation formula")
    parser.add_argument("--before", dest="before_group",
                        help="Insert before this group (default: before summary group or at end)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show what would be done without making changes")

    args = parser.parse_args()

    filepath = Path(args.xlsform)

    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    if not filepath.suffix.lower() == ".xlsx":
        print(f"WARNING: File does not have .xlsx extension: {filepath}")

    if args.dry_run:
        print(f"DRY RUN - Would add calculation to {filepath}:")
        print(f"  Field name: {args.field_name}")
        print(f"  Calculation: {args.calculation}")
        if args.before_group:
            print(f"  Insert before: {args.before_group}")
        sys.exit(0)

    result = add_calculation(filepath, args.field_name, args.calculation, args.before_group)

    if result["success"]:
        print(f"✓ Successfully added calculation to {result['file']}")
        print(f"  Field: {result['field_name']}")
        print(f"  Formula: {result['calculation']}")
        print(f"  Inserted at row: {result['inserted_at_row']}")
        print()
        print("Next steps:")
        print(f"  1. Review the change in {filepath}")
        print(f"  2. Run: cht --local convert-app-forms upload-app-forms -- {filepath.stem}")
    else:
        print(f"ERROR: {result['error']}")
        sys.exit(1)


if __name__ == "__main__":
    main()
