#!/usr/bin/env python3
"""
Read XLSForm files and extract structure for CHT task creation.
Uses openpyxl to parse Excel files.

Usage: python read-xlsform.py <path_to_xlsform.xlsx>
"""

import sys
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def read_sheet_as_dicts(workbook, sheet_name):
    """Read a sheet and return list of dicts with header keys."""
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    # First row is header
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

    result = []
    for row in rows[1:]:
        if not any(row):  # Skip empty rows
            continue
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers) and value is not None:
                row_dict[headers[i]] = str(value).strip() if value else ""
        if row_dict:
            result.append(row_dict)

    return result


def analyze_xlsform(filepath):
    """Analyze XLSForm and extract relevant information for task creation."""

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return {"error": f"Cannot open file: {e}"}

    result = {
        "file": str(filepath),
        "form_id": None,
        "form_title": None,
        "survey": {
            "fields": [],
            "groups": [],
            "date_fields": [],
            "calculated_fields": [],
            "select_fields": [],
            "important_fields": []
        },
        "choices": {},
        "settings": {}
    }

    # Read settings sheet
    settings = read_sheet_as_dicts(wb, "settings")
    if settings:
        result["settings"] = settings[0]
        result["form_id"] = settings[0].get("form_id", "")
        result["form_title"] = settings[0].get("form_title", "")

    # Read choices sheet
    choices = read_sheet_as_dicts(wb, "choices")
    for choice in choices:
        list_name = choice.get("list_name", "")
        if list_name:
            if list_name not in result["choices"]:
                result["choices"][list_name] = []
            result["choices"][list_name].append({
                "name": choice.get("name", ""),
                "label": choice.get("label", "")
            })

    # Read survey sheet
    survey = read_sheet_as_dicts(wb, "survey")

    # Important field patterns for task creation
    important_patterns = [
        "patient", "person", "_id", "date", "delivery", "birth",
        "risk", "danger", "referral", "follow", "visit", "next",
        "status", "outcome", "lmp", "edd", "pregnant", "assessment"
    ]

    current_group = None

    for row in survey:
        field_type = row.get("type", "").lower()
        field_name = row.get("name", "")
        field_label = row.get("label", "")
        relevant = row.get("relevant", "")
        calculation = row.get("calculation", "")

        # Track groups
        if field_type.startswith("begin"):
            if "group" in field_type or "repeat" in field_type:
                current_group = field_name
                result["survey"]["groups"].append({
                    "name": field_name,
                    "label": field_label,
                    "type": field_type
                })
        elif field_type.startswith("end"):
            current_group = None
            continue

        # Skip meta/structural types
        if field_type in ["end group", "end repeat", "end_group", "end_repeat", ""]:
            continue

        # Build field info
        field_info = {
            "name": field_name,
            "type": field_type,
            "label": field_label,
            "group": current_group
        }

        if relevant:
            field_info["relevant"] = relevant
        if calculation:
            field_info["calculation"] = calculation

        # Categorize fields
        if field_type in ["date", "datetime", "dateTime"]:
            result["survey"]["date_fields"].append(field_info)

        if field_type == "calculate":
            result["survey"]["calculated_fields"].append(field_info)

        if field_type.startswith("select"):
            list_name = field_type.split()[-1] if " " in field_type else ""
            field_info["choices_list"] = list_name
            result["survey"]["select_fields"].append(field_info)

        # Check if field matches important patterns
        field_lower = field_name.lower()
        for pattern in important_patterns:
            if pattern in field_lower:
                if field_info not in result["survey"]["important_fields"]:
                    result["survey"]["important_fields"].append(field_info)
                break

        # Add to all fields
        if field_name and field_type not in ["begin group", "begin_group", "begin repeat", "begin_repeat"]:
            result["survey"]["fields"].append(field_info)

    wb.close()
    return result


def format_output(analysis):
    """Format analysis for readable output."""

    if "error" in analysis:
        print(f"ERROR: {analysis['error']}")
        return

    print("=" * 60)
    print(f"XLSForm Analysis: {Path(analysis['file']).name}")
    print("=" * 60)
    print()

    # Form info
    print("=== Form Information ===")
    print(f"  Form ID: {analysis['form_id'] or '(not set)'}")
    print(f"  Title: {analysis['form_title'] or '(not set)'}")
    print()

    # Statistics
    print("=== Field Statistics ===")
    print(f"  Total fields: {len(analysis['survey']['fields'])}")
    print(f"  Groups: {len(analysis['survey']['groups'])}")
    print(f"  Date fields: {len(analysis['survey']['date_fields'])}")
    print(f"  Calculated fields: {len(analysis['survey']['calculated_fields'])}")
    print(f"  Select fields: {len(analysis['survey']['select_fields'])}")
    print(f"  Choice lists: {len(analysis['choices'])}")
    print()

    # Important fields for tasks
    print("=== Important Fields (for task triggers/conditions) ===")
    if analysis['survey']['important_fields']:
        for field in analysis['survey']['important_fields'][:15]:
            label = f" - {field['label']}" if field.get('label') else ""
            print(f"  {field['name']}: {field['type']}{label}")
        if len(analysis['survey']['important_fields']) > 15:
            print(f"  ... and {len(analysis['survey']['important_fields']) - 15} more")
    else:
        print("  (none detected)")
    print()

    # Date fields (key for scheduling)
    print("=== Date Fields (for task scheduling) ===")
    if analysis['survey']['date_fields']:
        for field in analysis['survey']['date_fields']:
            label = f" - {field['label']}" if field.get('label') else ""
            print(f"  {field['name']}{label}")
    else:
        print("  (none found)")
    print()

    # Select fields (for conditions)
    print("=== Select Fields (for task conditions) ===")
    if analysis['survey']['select_fields']:
        for field in analysis['survey']['select_fields'][:10]:
            choices_list = field.get('choices_list', '')
            choices = analysis['choices'].get(choices_list, [])
            choice_names = [c['name'] for c in choices[:5]]
            choices_str = ", ".join(choice_names)
            if len(choices) > 5:
                choices_str += f" ... (+{len(choices)-5})"
            print(f"  {field['name']}: [{choices_str}]")
        if len(analysis['survey']['select_fields']) > 10:
            print(f"  ... and {len(analysis['survey']['select_fields']) - 10} more")
    else:
        print("  (none found)")
    print()

    # Calculated fields
    print("=== Calculated Fields ===")
    if analysis['survey']['calculated_fields']:
        for field in analysis['survey']['calculated_fields'][:8]:
            calc = field.get('calculation', '')[:50]
            if len(field.get('calculation', '')) > 50:
                calc += "..."
            print(f"  {field['name']}: {calc}")
        if len(analysis['survey']['calculated_fields']) > 8:
            print(f"  ... and {len(analysis['survey']['calculated_fields']) - 8} more")
    else:
        print("  (none found)")
    print()

    # JSON output option
    print("=== JSON Output ===")
    print("(Use --json flag for machine-readable output)")


def main():
    if len(sys.argv) < 2:
        print("Usage: python read-xlsform.py <path_to_xlsform.xlsx> [--json]")
        print()
        print("Analyzes an XLSForm file and extracts field information")
        print("useful for creating CHT tasks.")
        sys.exit(1)

    filepath = sys.argv[1]
    json_output = "--json" in sys.argv

    if not Path(filepath).exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    if not filepath.endswith(".xlsx"):
        print(f"WARNING: File does not have .xlsx extension: {filepath}")

    analysis = analyze_xlsform(filepath)

    if json_output:
        print(json.dumps(analysis, indent=2))
    else:
        format_output(analysis)


if __name__ == "__main__":
    main()
