# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl", "unidecode"]
# ///
"""
Parse CHT form design Excel documents into structured JSON.

Handles real-world CIV design format with:
- Two Option columns (F & G)
- Multi-row option continuations
- Session-level conditions (condition in Type or Question column)
- Template variables like {{Prenom_du_patient}}
- Image URLs in columns after Observations
- Multi-line conditions preserved as-is
- Nested output: groups -> sessions -> fields

Usage:
    uv run scripts/parse-design-sheet.py <file_path>                     # list sheets
    uv run scripts/parse-design-sheet.py <file_path> "PEC - 5 ans"       # parse sheet
    uv run scripts/parse-design-sheet.py <file_path> "PEC - 5 ans" --pretty
"""

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from unidecode import unidecode


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Fixed column indices for the CIV design format (0-based)
COL_GROUP = 0       # A: Groupe Label
COL_SESSION = 1     # B: Session
COL_QUESTION = 2    # C: Question
COL_TYPE = 3        # D: Type
COL_REQUIRED = 4    # E: Obligatoire
COL_OPTION1 = 5     # F: Option (first)
COL_OPTION2 = 6     # G: Option (second)
COL_CONDITION = 7   # H: Condition
COL_OBSERVATIONS = 8  # I: Observations
COL_IMAGES_START = 9  # J onwards: image URLs

# Type normalization map
TYPE_MAP: dict[str, str] = {
    "texte": "text",
    "text": "text",
    "nombre": "integer",
    "number": "integer",
    "integer": "integer",
    "entier": "integer",
    "decimal": "decimal",
    "décimal": "decimal",
    "date": "date",
    "time": "time",
    "heure": "time",
    "datetime": "dateTime",
    "select_one": "select_one",
    "select one": "select_one",
    "choix unique": "select_one",
    "liste déroulante": "select_one",
    "select_multiple": "select_multiple",
    "select multiple": "select_multiple",
    "choix multiple": "select_multiple",
    "checkbox": "select_multiple",
    "note": "note",
    "calculate": "calculate",
    "calcul": "calculate",
    "geopoint": "geopoint",
    "geolocation": "geopoint",
    "géolocalisation": "geopoint",
    "image": "image",
    "photo": "image",
    "barcode": "barcode",
    "code-barres": "barcode",
    "phone": "string",
    "téléphone": "string",
    "email": "string",
    "textarea": "text",
    "zone de texte": "text",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def cell_str(value: Any) -> str | None:
    """Return stripped string or None for empty/whitespace-only values."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def is_yes(value: Any) -> bool:
    """Check if a value represents 'yes' in French or English."""
    s = cell_str(value)
    if not s:
        return False
    return s.lower() in ("oui", "yes", "o", "y", "1", "true", "x")


def normalize_type(raw: str | None) -> str:
    """Normalize a type string to a canonical XLSForm type."""
    if not raw:
        return "text"
    key = raw.lower().strip()
    return TYPE_MAP.get(key, key)


def slugify(text: str, strip_templates: bool = True) -> str:
    """
    Convert French text to a valid XLSForm field name.

    - Optionally strips {{template_vars}} before slugifying
    - Uses unidecode for accent removal
    - Lowercase, non-alnum -> _, collapse multiples
    - Prefix q_ if starts with digit
    - Max 50 chars
    """
    if not text:
        return ""

    s = str(text)

    # Strip template variables like {{Prenom_du_patient}}
    if strip_templates:
        s = re.sub(r"\{\{[^}]*\}\}", "", s)

    # Convert to ASCII
    s = unidecode(s)
    s = s.lower()

    # Replace non-alphanumeric with underscore
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    s = re.sub(r"_+", "_", s)

    # Prefix if starts with digit
    if s and s[0].isdigit():
        s = "q_" + s

    # Truncate
    if len(s) > 50:
        s = s[:50].rstrip("_")

    return s or "unnamed"


def is_condition_text(text: str) -> bool:
    """Heuristic: does this text look like a session-level condition?"""
    t = text.strip().lower()
    return t.startswith("si ") or t.startswith("si\n")


def extract_image_urls(row_data: list) -> list[str]:
    """Extract Google Drive / image URLs from columns J onwards."""
    urls = []
    for idx in range(COL_IMAGES_START, len(row_data)):
        val = cell_str(row_data[idx])
        if val and ("drive.google.com" in val or "http" in val.lower()):
            urls.append(val)
    return urls


def row_is_empty(row_data: list) -> bool:
    """True if every cell in the row is None or whitespace."""
    return all(cell_str(v) is None for v in row_data)


# ---------------------------------------------------------------------------
# Name uniqueness tracker
# ---------------------------------------------------------------------------

class NameTracker:
    """Generate unique slugified names, appending _2, _3... for duplicates."""

    def __init__(self) -> None:
        self._counts: dict[str, int] = {}

    def unique(self, label: str) -> str:
        base = slugify(label)
        if not base:
            base = "unnamed"
        if base not in self._counts:
            self._counts[base] = 1
            return base
        self._counts[base] += 1
        return f"{base}_{self._counts[base]}"


# ---------------------------------------------------------------------------
# Choice list manager
# ---------------------------------------------------------------------------

class ChoiceListManager:
    """Manage choice lists with deduplication and yes/no detection."""

    YES_NO_LABELS = ({"oui", "non"}, {"yes", "no"}, {"vrai", "faux"})

    def __init__(self) -> None:
        self.lists: dict[str, list[dict]] = {}

    def _labels_set(self, options: list[dict]) -> frozenset[str]:
        return frozenset(o["label"]["fr"].lower().strip() for o in options)

    def _is_yes_no(self, options: list[dict]) -> bool:
        if len(options) != 2:
            return False
        labels = self._labels_set(options)
        return labels in self.YES_NO_LABELS

    def _build_options(self, raw_labels: list[str]) -> list[dict]:
        """Turn a list of raw label strings into structured option dicts."""
        opts = []
        seen_names: dict[str, int] = {}
        for label in raw_labels:
            label = label.strip()
            if not label:
                continue
            name = slugify(label)
            if not name:
                name = "opt"
            # Handle duplicate option names within the same list
            if name in seen_names:
                seen_names[name] += 1
                name = f"{name}_{seen_names[name]}"
            else:
                seen_names[name] = 1
            opts.append({"name": name, "label": {"fr": label}})
        return opts

    def register(self, field_name: str, raw_labels: list[str]) -> str | None:
        """
        Register options for a field. Returns the list name to use,
        or None if no valid options.
        """
        options = self._build_options(raw_labels)
        if not options:
            return None

        # Check yes/no
        if self._is_yes_no(options):
            if "yes_no" not in self.lists:
                self.lists["yes_no"] = [
                    {"name": "yes", "label": {"fr": "Oui"}},
                    {"name": "no", "label": {"fr": "Non"}},
                ]
            return "yes_no"

        # Check for existing identical list
        new_labels = self._labels_set(options)
        for existing_name, existing_opts in self.lists.items():
            if self._labels_set(existing_opts) == new_labels:
                return existing_name

        # Create new list
        list_name = f"{field_name}_list"
        # Ensure list name is unique
        suffix = 2
        base_list_name = list_name
        while list_name in self.lists:
            list_name = f"{base_list_name}_{suffix}"
            suffix += 1
        self.lists[list_name] = options
        return list_name

    def to_dict(self) -> dict:
        return dict(self.lists)


# ---------------------------------------------------------------------------
# Row classification
# ---------------------------------------------------------------------------

def classify_row(row_data: list) -> str:
    """
    Classify what kind of row this is:
    - 'empty'
    - 'group_session': has group and/or session label
    - 'session_with_condition': session row where Type column has a condition
    - 'question': has question text
    - 'option_continuation': no question but has option values
    """
    if row_is_empty(row_data):
        return "empty"

    group = cell_str(row_data[COL_GROUP]) if len(row_data) > COL_GROUP else None
    session = cell_str(row_data[COL_SESSION]) if len(row_data) > COL_SESSION else None
    question = cell_str(row_data[COL_QUESTION]) if len(row_data) > COL_QUESTION else None
    type_val = cell_str(row_data[COL_TYPE]) if len(row_data) > COL_TYPE else None
    opt1 = cell_str(row_data[COL_OPTION1]) if len(row_data) > COL_OPTION1 else None
    opt2 = cell_str(row_data[COL_OPTION2]) if len(row_data) > COL_OPTION2 else None

    # Session row with condition in Type column
    # Pattern: session is set, no question text, and Type looks like a condition
    # Also handle: session set, condition in Question column (e.g. row 123)
    if session and not question and type_val and is_condition_text(type_val):
        return "session_with_condition"
    if session and question and not type_val and is_condition_text(question):
        return "session_with_condition_in_question"

    # Group or session header (may also have a question on the same row)
    if question:
        return "question"

    # No question but has options -> continuation
    if opt1 or opt2:
        return "option_continuation"

    # Session or group label only (no question, no condition in type)
    if session or group:
        return "group_session"

    return "empty"


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def get_cell(row_data: list, idx: int) -> str | None:
    """Safely get a cell value from row_data."""
    if idx < len(row_data):
        return cell_str(row_data[idx])
    return None


def collect_row_options(row_data: list) -> list[str]:
    """Collect non-empty option values from both Option columns."""
    opts = []
    for col in (COL_OPTION1, COL_OPTION2):
        val = get_cell(row_data, col)
        if val:
            opts.append(val)
    return opts


def detect_header_row(ws) -> int:
    """Find the header row index (1-based). Returns 1 if standard format detected."""
    for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), start=1):
        vals = [cell_str(v) for v in row]
        # Look for characteristic header columns
        lower_vals = [v.lower() if v else "" for v in vals]
        if any("question" in v for v in lower_vals) or any("type" in v for v in lower_vals):
            return row_idx
    return 1  # Default to row 1


def parse_sheet(ws, sheet_name: str, source_file: str) -> dict:
    """
    Parse a worksheet into structured nested JSON.

    Returns the full output structure with groups, sessions, fields, choice_lists, summary.
    """
    names = NameTracker()
    choices = ChoiceListManager()

    # Result structure
    groups: list[dict] = []
    current_group: dict | None = None
    current_session: dict | None = None
    current_field: dict | None = None  # Field being built (may accumulate options)

    # Track state
    current_group_label: str | None = None
    current_session_label: str | None = None

    # Field type counts for summary
    field_type_counts: dict[str, int] = {}
    total_conditions = 0

    header_row = detect_header_row(ws)

    def ensure_group(label: str | None = None) -> dict:
        """Ensure there is a current group, creating a default if needed."""
        nonlocal current_group, current_group_label
        if label:
            current_group_label = label
        if current_group is None or (label and label != current_group.get("label", {}).get("fr")):
            grp_label = label or current_group_label or "(Sans groupe)"
            current_group = {
                "name": names.unique(grp_label),
                "label": {"fr": grp_label},
                "sessions": [],
            }
            groups.append(current_group)
        return current_group

    def ensure_session(label: str | None = None, condition: str | None = None) -> dict:
        """Ensure there is a current session."""
        nonlocal current_session, current_session_label
        if label:
            current_session_label = label
        if current_session is None or (label and label != current_session.get("label", {}).get("fr")):
            sess_label = label or current_session_label or "(Sans session)"
            current_session = {
                "name": names.unique(sess_label),
                "label": {"fr": sess_label},
                "condition": condition,
                "fields": [],
            }
            grp = ensure_group()
            grp["sessions"].append(current_session)
        elif condition and current_session.get("condition") is None:
            current_session["condition"] = condition
        return current_session

    def finalize_field():
        """Finalize the current field: process accumulated options and add to session."""
        nonlocal current_field, total_conditions
        if current_field is None:
            return

        field = current_field
        current_field = None

        raw_opts = field.pop("_raw_options", [])
        field_type = field["type"]

        # Process options for select types
        if raw_opts and field_type in ("select_one", "select_multiple"):
            list_name = choices.register(field["name"], raw_opts)
            field["options"] = list_name
        elif raw_opts:
            # Field has options but type is not select -- might still need them
            # (e.g., note with options in some edge cases)
            list_name = choices.register(field["name"], raw_opts)
            if list_name:
                field["options"] = list_name

        if "options" not in field:
            field["options"] = None

        # Count conditions
        if field.get("condition"):
            total_conditions += 1

        # Track type counts
        ft = field["type"]
        field_type_counts[ft] = field_type_counts.get(ft, 0) + 1

        # Add to current session
        sess = ensure_session()
        sess["fields"].append(field)

    # ----- Process rows -----
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_data = list(row)
        kind = classify_row(row_data)

        if kind == "empty":
            continue

        # --- Update group if present ---
        group_label = get_cell(row_data, COL_GROUP)
        if group_label:
            # Finalize any pending field before switching group
            finalize_field()
            current_session = None
            ensure_group(group_label)

        # --- Session with condition in Type column ---
        if kind == "session_with_condition":
            finalize_field()
            session_label = get_cell(row_data, COL_SESSION)
            condition = get_cell(row_data, COL_TYPE)
            current_session = None
            ensure_session(session_label, condition)
            continue

        # --- Session with condition in Question column (e.g. row 123) ---
        if kind == "session_with_condition_in_question":
            finalize_field()
            session_label = get_cell(row_data, COL_SESSION)
            condition = get_cell(row_data, COL_QUESTION)
            current_session = None
            ensure_session(session_label, condition)
            continue

        # --- Group/session header without question ---
        if kind == "group_session":
            finalize_field()
            session_label = get_cell(row_data, COL_SESSION)
            if session_label:
                current_session = None
                ensure_session(session_label)
            continue

        # --- Option continuation row ---
        if kind == "option_continuation":
            opts = collect_row_options(row_data)
            # Also collect image URLs from continuation rows
            img_urls = extract_image_urls(row_data)
            if current_field:
                current_field["_raw_options"].extend(opts)
                if img_urls:
                    current_field["image_urls"].extend(img_urls)
            continue

        # --- Question row ---
        if kind == "question":
            # Finalize previous field
            finalize_field()

            # Update session if session column has value
            session_label = get_cell(row_data, COL_SESSION)
            if session_label:
                current_session = None
                ensure_session(session_label)

            question_text = get_cell(row_data, COL_QUESTION) or ""
            raw_type = get_cell(row_data, COL_TYPE)
            field_type = normalize_type(raw_type)
            required = is_yes(get_cell(row_data, COL_REQUIRED))
            condition = get_cell(row_data, COL_CONDITION)
            observations = get_cell(row_data, COL_OBSERVATIONS)
            image_urls = extract_image_urls(row_data)

            # Collect initial options from both columns
            raw_opts = collect_row_options(row_data)

            # Generate unique name
            field_name = names.unique(question_text)

            # Parse observations for constraints vs hints
            constraint = None
            hint = None
            if observations:
                obs_lower = observations.lower()
                if any(kw in obs_lower for kw in [">=", "<=", ">", "<", "entre", "between", "compris", "min", "max"]):
                    constraint = observations
                elif not obs_lower.startswith("ins"):  # Skip "Insérer les images" type notes
                    hint = observations

            current_field = {
                "name": field_name,
                "type": field_type,
                "label": {"fr": question_text},
                "required": required,
                "condition": condition,
                "constraint": constraint,
                "hint": hint,
                "image_urls": image_urls,
                "observations": observations,
                "_raw_options": raw_opts,
            }

    # Finalize last field
    finalize_field()

    # Prune empty groups and sessions
    for grp in groups:
        grp["sessions"] = [s for s in grp["sessions"] if s["fields"]]
    groups = [g for g in groups if g["sessions"]]

    # Count session-level conditions
    for grp in groups:
        for sess in grp["sessions"]:
            if sess.get("condition"):
                total_conditions += 1

    # Build summary
    total_fields = sum(
        len(sess["fields"])
        for grp in groups
        for sess in grp["sessions"]
    )
    total_sessions = sum(len(grp["sessions"]) for grp in groups)

    return {
        "source_file": source_file,
        "source_sheet": sheet_name,
        "groups": groups,
        "choice_lists": choices.to_dict(),
        "summary": {
            "total_fields": total_fields,
            "total_groups": len(groups),
            "total_sessions": total_sessions,
            "total_conditions": total_conditions,
            "total_choice_lists": len(choices.lists),
            "field_types": dict(sorted(field_type_counts.items())),
        },
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Parse CHT form design Excel documents into structured JSON",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    uv run scripts/parse-design-sheet.py design.xlsx
    uv run scripts/parse-design-sheet.py design.xlsx "PEC - 5 ans"
    uv run scripts/parse-design-sheet.py design.xlsx "PEC - 5 ans" --pretty
        """,
    )
    parser.add_argument("file_path", help="Path to Excel file")
    parser.add_argument("sheet_name", nargs="?", help="Sheet name to parse")
    parser.add_argument(
        "--pretty", "-p", action="store_true", help="Pretty-print JSON output"
    )
    parser.add_argument("--output", "-o", help="Output file path (default: stdout)")
    parser.add_argument(
        "--list-sheets", action="store_true",
        help="List available sheets and exit (same as omitting sheet_name)"
    )

    args = parser.parse_args()
    if args.list_sheets:
        args.sheet_name = None  # Force sheet listing mode
    file_path = Path(args.file_path)

    if not file_path.exists():
        print(f"Error: file not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}", file=sys.stderr)
        sys.exit(1)

    # List sheets if no sheet name given
    if not args.sheet_name:
        print("Available sheets:", file=sys.stderr)
        for name in wb.sheetnames:
            print(f"  - {name}", file=sys.stderr)
        print(
            f"\nUsage: {sys.argv[0]} {args.file_path} <sheet_name>",
            file=sys.stderr,
        )
        sys.exit(0)

    if args.sheet_name not in wb.sheetnames:
        print(f"Error: sheet '{args.sheet_name}' not found.", file=sys.stderr)
        print("Available sheets:", file=sys.stderr)
        for name in wb.sheetnames:
            print(f"  - {name}", file=sys.stderr)
        sys.exit(1)

    ws = wb[args.sheet_name]
    result = parse_sheet(ws, args.sheet_name, file_path.name)

    # Output
    indent = 2 if args.pretty else None
    json_output = json.dumps(result, ensure_ascii=False, indent=indent)

    if args.output:
        Path(args.output).write_text(json_output, encoding="utf-8")
        print(f"Written to {args.output}", file=sys.stderr)
    else:
        print(json_output)

    # Print summary to stderr
    s = result["summary"]
    print(
        f"\n--- Summary ---\n"
        f"Groups:       {s['total_groups']}\n"
        f"Sessions:     {s['total_sessions']}\n"
        f"Fields:       {s['total_fields']}\n"
        f"Conditions:   {s['total_conditions']}\n"
        f"Choice lists: {s['total_choice_lists']}\n"
        f"Field types:  {s['field_types']}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
