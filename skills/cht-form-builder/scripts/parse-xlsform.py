#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""
XLSForm Parser - Parse XLSForm Excel files and output structured JSON.

This script reads an XLSForm Excel file and extracts:
- Survey questions with hierarchy and groups
- Choice lists
- Form settings and metadata

Usage:
    uv run parse-xlsform.py <xlsform_path>

Output:
    JSON structure to stdout
"""

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def detect_languages(headers: list[str]) -> list[str]:
    """
    Detect languages from column headers like label::fr, label::en, hint::sw.
    Returns a list of unique language codes found.
    """
    languages = set()
    pattern = re.compile(r'^(?:label|hint|constraint_message|required_message)::(\w+)$')

    for header in headers:
        if header:
            match = pattern.match(str(header).strip().lower())
            if match:
                languages.add(match.group(1))

    return sorted(languages) if languages else []


def get_column_index(headers: list[str], *column_names: str) -> int | None:
    """
    Find the index of a column by trying multiple possible names.
    Returns None if not found.
    """
    headers_lower = [str(h).strip().lower() if h else '' for h in headers]
    for name in column_names:
        name_lower = name.lower()
        if name_lower in headers_lower:
            return headers_lower.index(name_lower)
    return None


def get_localized_values(row: list[Any], headers: list[str], base_name: str, languages: list[str]) -> dict[str, str]:
    """
    Extract localized values for a field (e.g., label::fr, label::en).
    Returns a dict with language codes as keys.
    """
    values = {}
    headers_lower = [str(h).strip().lower() if h else '' for h in headers]

    # First try language-specific columns
    for lang in languages:
        col_name = f"{base_name}::{lang}"
        if col_name in headers_lower:
            idx = headers_lower.index(col_name)
            if idx < len(row) and row[idx]:
                values[lang] = str(row[idx]).strip()

    # If no language-specific values found, try the base column
    if not values:
        base_idx = get_column_index(headers, base_name)
        if base_idx is not None and base_idx < len(row) and row[base_idx]:
            # Use 'default' as key when no language specified
            values['default'] = str(row[base_idx]).strip()

    return values


def parse_survey_sheet(ws, languages: list[str]) -> tuple[list[dict], list[str]]:
    """
    Parse the survey sheet and extract questions with hierarchy.
    Returns a tuple of (questions list, groups list).
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(h).strip().lower() if h else '' for h in rows[0]]

    # Column indices
    type_idx = get_column_index(headers, 'type')
    name_idx = get_column_index(headers, 'name')
    required_idx = get_column_index(headers, 'required')
    relevant_idx = get_column_index(headers, 'relevant')
    appearance_idx = get_column_index(headers, 'appearance')
    constraint_idx = get_column_index(headers, 'constraint')
    calculation_idx = get_column_index(headers, 'calculation', 'calculate')
    readonly_idx = get_column_index(headers, 'readonly', 'read_only')
    default_idx = get_column_index(headers, 'default')
    repeat_count_idx = get_column_index(headers, 'repeat_count')

    questions = []
    groups = []
    group_stack = []  # Track nested groups

    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        # Get type and name
        q_type = str(row[type_idx]).strip() if type_idx is not None and row[type_idx] else ''
        q_name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else ''

        if not q_type:
            continue

        q_type_lower = q_type.lower()

        # Handle group/repeat start
        if q_type_lower.startswith('begin group') or q_type_lower.startswith('begin_group'):
            group_stack.append(q_name)
            if q_name and q_name not in groups:
                groups.append(q_name)
            continue

        # Handle repeat start
        if q_type_lower.startswith('begin repeat') or q_type_lower.startswith('begin_repeat'):
            group_stack.append(q_name)
            if q_name and q_name not in groups:
                groups.append(q_name)
            continue

        # Handle group/repeat end
        if q_type_lower.startswith('end group') or q_type_lower.startswith('end_group'):
            if group_stack:
                group_stack.pop()
            continue

        if q_type_lower.startswith('end repeat') or q_type_lower.startswith('end_repeat'):
            if group_stack:
                group_stack.pop()
            continue

        # Skip notes and other non-question types if they have no name
        if not q_name:
            continue

        # Build question object
        question = {
            'name': q_name,
            'type': q_type,
            'labels': get_localized_values(row, headers, 'label', languages),
        }

        # Add hints if present
        hints = get_localized_values(row, headers, 'hint', languages)
        if hints:
            question['hints'] = hints

        # Required field
        if required_idx is not None and row[required_idx]:
            req_val = str(row[required_idx]).strip().lower()
            question['required'] = req_val in ('yes', 'true', '1', 'oui')
        else:
            question['required'] = False

        # Relevant condition
        if relevant_idx is not None and row[relevant_idx]:
            question['relevant'] = str(row[relevant_idx]).strip()

        # Appearance
        if appearance_idx is not None and row[appearance_idx]:
            question['appearance'] = str(row[appearance_idx]).strip()

        # Constraint
        if constraint_idx is not None and row[constraint_idx]:
            question['constraint'] = str(row[constraint_idx]).strip()
            # Also get constraint message
            constraint_msgs = get_localized_values(row, headers, 'constraint_message', languages)
            if constraint_msgs:
                question['constraint_message'] = constraint_msgs

        # Calculation
        if calculation_idx is not None and row[calculation_idx]:
            question['calculation'] = str(row[calculation_idx]).strip()

        # Read-only
        if readonly_idx is not None and row[readonly_idx]:
            ro_val = str(row[readonly_idx]).strip().lower()
            if ro_val in ('yes', 'true', '1', 'oui'):
                question['readonly'] = True

        # Default value
        if default_idx is not None and row[default_idx]:
            question['default'] = str(row[default_idx]).strip()

        # Current group path
        if group_stack:
            question['group'] = '/'.join(group_stack)

        # Handle select types - extract choice list name
        if q_type_lower.startswith('select_one ') or q_type_lower.startswith('select_multiple '):
            parts = q_type.split(None, 1)
            if len(parts) > 1:
                choice_list = parts[1].strip()
                # Remove any or_other suffix
                if choice_list.endswith(' or_other'):
                    choice_list = choice_list.rsplit(' ', 1)[0]
                question['choices'] = choice_list
                question['type'] = parts[0]  # Normalize type

        questions.append(question)

    return questions, groups


def parse_choices_sheet(ws, languages: list[str]) -> dict[str, list[dict]]:
    """
    Parse the choices sheet and extract choice lists grouped by list_name.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(h).strip().lower() if h else '' for h in rows[0]]

    list_name_idx = get_column_index(headers, 'list_name', 'list name')
    name_idx = get_column_index(headers, 'name')

    if list_name_idx is None or name_idx is None:
        return {}

    choice_lists = {}

    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        list_name = str(row[list_name_idx]).strip() if row[list_name_idx] else ''
        choice_name = str(row[name_idx]).strip() if row[name_idx] else ''

        if not list_name or not choice_name:
            continue

        choice = {
            'name': choice_name,
            'labels': get_localized_values(row, headers, 'label', languages),
        }

        if list_name not in choice_lists:
            choice_lists[list_name] = []

        choice_lists[list_name].append(choice)

    return choice_lists


def parse_settings_sheet(ws) -> dict[str, Any]:
    """
    Parse the settings sheet and extract form metadata.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    values = rows[1] if len(rows) > 1 else []

    settings = {}

    # Common settings fields
    field_mappings = {
        'form_id': ['form_id', 'id_string'],
        'form_title': ['form_title', 'title'],
        'version': ['version'],
        'default_language': ['default_language'],
        'style': ['style'],
        'instance_name': ['instance_name'],
    }

    for key, possible_names in field_mappings.items():
        idx = get_column_index(headers, *possible_names)
        if idx is not None and idx < len(values) and values[idx]:
            settings[key] = str(values[idx]).strip()

    return settings


def parse_xlsform(filepath: str) -> dict[str, Any]:
    """
    Parse an XLSForm Excel file and return structured data.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    if not path.suffix.lower() in ('.xlsx', '.xls'):
        raise ValueError(f"File must be an Excel file (.xlsx or .xls): {filepath}")

    wb = load_workbook(filename=filepath, read_only=True, data_only=True)

    result = {
        'form_id': None,
        'form_title': None,
        'languages': [],
        'questions': [],
        'choice_lists': {},
        'groups': [],
    }

    # Get sheet names (case-insensitive)
    sheet_names = {name.lower(): name for name in wb.sheetnames}

    # Detect languages from survey sheet headers first
    languages = []
    if 'survey' in sheet_names:
        survey_ws = wb[sheet_names['survey']]
        first_row = next(survey_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row:
            languages = detect_languages(list(first_row))

    # Also check choices sheet for additional languages
    if 'choices' in sheet_names:
        choices_ws = wb[sheet_names['choices']]
        first_row = next(choices_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row:
            choice_languages = detect_languages(list(first_row))
            for lang in choice_languages:
                if lang not in languages:
                    languages.append(lang)

    result['languages'] = sorted(languages) if languages else []

    # Parse settings sheet
    if 'settings' in sheet_names:
        settings = parse_settings_sheet(wb[sheet_names['settings']])
        result['form_id'] = settings.get('form_id')
        result['form_title'] = settings.get('form_title')
        if settings.get('default_language'):
            result['default_language'] = settings['default_language']
        if settings.get('version'):
            result['version'] = settings['version']

    # Parse survey sheet
    if 'survey' in sheet_names:
        questions, groups = parse_survey_sheet(wb[sheet_names['survey']], languages)
        result['questions'] = questions
        result['groups'] = groups

    # Parse choices sheet
    if 'choices' in sheet_names:
        result['choice_lists'] = parse_choices_sheet(wb[sheet_names['choices']], languages)

    wb.close()

    return result


def main():
    parser = argparse.ArgumentParser(
        description='Parse XLSForm Excel files and output structured JSON.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example usage:
    python parse-xlsform.py form.xlsx > output.json
    python parse-xlsform.py form.xlsx | jq '.questions'
        """
    )
    parser.add_argument(
        'xlsform_path',
        help='Path to the XLSForm Excel file (.xlsx or .xls)'
    )
    parser.add_argument(
        '--pretty', '-p',
        action='store_true',
        help='Pretty-print JSON output with indentation'
    )

    args = parser.parse_args()

    try:
        result = parse_xlsform(args.xlsform_path)

        if args.pretty:
            print(json.dumps(result, indent=2, ensure_ascii=False))
        else:
            print(json.dumps(result, ensure_ascii=False))

    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error parsing XLSForm: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
