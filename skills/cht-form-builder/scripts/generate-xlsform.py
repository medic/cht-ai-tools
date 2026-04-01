# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""
Generate CHT XLSForm Excel files from processed design JSON.

Reads the nested JSON output from parse-design-sheet.py (groups -> sessions -> fields)
and generates a complete XLSForm with survey, choices, and settings sheets that conform
to the CHT application framework conventions.

Usage:
    uv run scripts/generate-xlsform.py input.json output.xlsx
    uv run scripts/generate-xlsform.py input.json output.xlsx --languages fr,en
    uv run scripts/generate-xlsform.py input.json output.xlsx --form-id patient_assessment_under_5
    uv run scripts/generate-xlsform.py input.json output.xlsx --form-title "Patient Assessment"
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Standard CHT calculate fields: (name, calculation)
STANDARD_CALCULATES = [
    ("patient_id", "../inputs/contact/_id"),
    ("household_id", "../inputs/contact/parent/_id"),
    ("patient_first_name", "../inputs/contact/first_name"),
    ("patient_last_name", "../inputs/contact/last_name"),
    ("patient_full_name", "concat(${patient_first_name}, ' ', ${patient_last_name})"),
    ("patient_name", "coalesce(../inputs/contact/name, ${patient_full_name})"),
    ("date_of_birth", "../inputs/contact/date_of_birth"),
    ("patient_age_in_years", "floor(difference-in-months(${date_of_birth}, today()) div 12)"),
    ("patient_age_in_months", "difference-in-months(${date_of_birth}, now())"),
    ("patient_age_in_days", "floor(decimal-date-time(today()) - decimal-date-time(${date_of_birth}))"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def label_for_langs(labels: dict | str | None, languages: list[str]) -> dict[str, str]:
    """Build {label::fr: ..., label::en: ...} dict from a label value.

    Handles:
    - dict with language keys: {"fr": "Bonjour", "en": "Hello"}
    - plain string: applied to all languages
    - None/empty: empty string for all languages
    """
    result = {}
    if isinstance(labels, dict):
        for lang in languages:
            result[f"label::{lang}"] = labels.get(lang, "")
    elif isinstance(labels, str) and labels:
        for lang in languages:
            result[f"label::{lang}"] = labels
    else:
        for lang in languages:
            result[f"label::{lang}"] = ""
    return result


def no_label(languages: list[str]) -> dict[str, str]:
    """Return NO_LABEL for all languages."""
    return {f"label::{lang}": "NO_LABEL" for lang in languages}


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

def survey_columns(languages: list[str]) -> list[str]:
    """Build ordered column names for the survey sheet."""
    cols = ["type", "name"]
    for lang in languages:
        cols.append(f"label::{lang}")
    cols.extend(["required", "relevant", "appearance", "constraint"])
    for lang in languages:
        cols.append(f"constraint_message::{lang}")
    cols.append("calculation")
    cols.append("default")
    for lang in languages:
        cols.append(f"hint::{lang}")
    return cols


def choices_columns(languages: list[str]) -> list[str]:
    cols = ["list_name", "name"]
    for lang in languages:
        cols.append(f"label::{lang}")
    return cols


def settings_columns() -> list[str]:
    return ["form_title", "form_id", "version", "style"]


# ---------------------------------------------------------------------------
# Survey row builders
# ---------------------------------------------------------------------------

def build_inputs_group(languages: list[str]) -> list[dict]:
    """Build the full CHT inputs group (contact resolution boilerplate)."""
    rows: list[dict] = []

    # begin group inputs
    row = {"type": "begin group", "name": "inputs",
           "relevant": "./source = 'user'",
           "appearance": "field-list summary"}
    row.update(label_for_langs("Patient", languages))
    rows.append(row)

    # hidden source
    rows.append({"type": "hidden", "name": "source"})

    # hidden source_id
    rows.append({"type": "hidden", "name": "source_id"})

    # begin group user
    row = {"type": "begin group", "name": "user"}
    row.update(no_label(languages))
    rows.append(row)

    # db:person contact_id
    row = {"type": "db:person", "name": "contact_id", "appearance": "db-object"}
    row.update(no_label(languages))
    rows.append(row)

    # end group user
    rows.append({"type": "end group", "name": "user"})

    # begin group contact
    row = {"type": "begin group", "name": "contact"}
    row.update(no_label(languages))
    rows.append(row)

    # db:person _id (required)
    row = {"type": "db:person", "name": "_id", "required": "yes",
           "appearance": "db-object"}
    row.update(no_label(languages))
    rows.append(row)

    # hidden patient_id
    rows.append({"type": "hidden", "name": "patient_id"})

    # string fields for contact data
    for field_name in ["date_of_birth", "first_name", "last_name"]:
        row = {"type": "string", "name": field_name}
        row.update(no_label(languages))
        rows.append(row)

    # name with hidden appearance
    row = {"type": "string", "name": "name", "appearance": "hidden"}
    row.update(no_label(languages))
    rows.append(row)

    # sex
    row = {"type": "string", "name": "sex"}
    row.update(no_label(languages))
    rows.append(row)

    # begin group parent (hidden)
    row = {"type": "begin group", "name": "parent", "appearance": "hidden"}
    row.update(no_label(languages))
    rows.append(row)

    # parent _id
    row = {"type": "string", "name": "_id"}
    row.update(no_label(languages))
    rows.append(row)

    # end group parent
    rows.append({"type": "end group", "name": "parent"})

    # end group contact
    rows.append({"type": "end group", "name": "contact"})

    # hidden task_id
    rows.append({"type": "hidden", "name": "task_id"})

    # end group inputs
    rows.append({"type": "end group", "name": "inputs"})

    return rows


def build_standard_calculates() -> list[dict]:
    """Build standard patient calculate fields."""
    return [{"type": "calculate", "name": name, "calculation": calc}
            for name, calc in STANDARD_CALCULATES]


def build_contact_summary_calculates(var_names: list[str]) -> list[dict]:
    """Build calculate rows that pull from contact-summary context."""
    return [
        {"type": "calculate", "name": var,
         "calculation": f"instance('contact-summary')/context/{var}"}
        for var in var_names
    ]


def build_inferred_calculates(entries: list[dict]) -> list[dict]:
    """Build calculate rows for inferred/derived values.

    Each entry: {"name": "xxx", "calculation": "...xpath..."}
    """
    return [
        {"type": "calculate", "name": e["name"], "calculation": e["calculation"]}
        for e in entries
    ]


def build_field_row(field: dict, languages: list[str]) -> dict:
    """Convert a single design field to a survey row dict."""
    row: dict = {}

    field_type = field.get("type", "text")

    # Handle select types: attach list name
    if field_type in ("select_one", "select_multiple"):
        list_ref = field.get("options") or field.get("list_name") or field.get("name", "choices")
        row["type"] = f"{field_type} {list_ref}"
    else:
        row["type"] = field_type

    row["name"] = field.get("name", "")

    # Labels (parser outputs {"fr": "..."} in field["label"])
    label = field.get("label")
    row.update(label_for_langs(label, languages))

    # Required
    if field.get("required"):
        row["required"] = "yes"

    # Relevant (skip logic) -- may come from field["relevant"] or field["condition"]
    relevant = field.get("relevant") or field.get("condition")
    if relevant:
        row["relevant"] = relevant

    # Appearance
    if field.get("appearance"):
        row["appearance"] = field["appearance"]

    # Constraint
    if field.get("constraint"):
        row["constraint"] = field["constraint"]

    # Constraint messages
    constraint_msg = field.get("constraint_message")
    if isinstance(constraint_msg, dict):
        for lang in languages:
            if lang in constraint_msg:
                row[f"constraint_message::{lang}"] = constraint_msg[lang]
    elif isinstance(constraint_msg, str) and constraint_msg:
        for lang in languages:
            row[f"constraint_message::{lang}"] = constraint_msg

    # Calculation
    if field.get("calculation"):
        row["calculation"] = field["calculation"]

    # Default
    if field.get("default"):
        row["default"] = field["default"]

    # Hints
    hint = field.get("hint")
    if isinstance(hint, dict):
        for lang in languages:
            if lang in hint:
                row[f"hint::{lang}"] = hint[lang]
    elif isinstance(hint, str) and hint:
        for lang in languages:
            row[f"hint::{lang}"] = hint

    return row


def build_session_rows(session: dict, languages: list[str]) -> list[dict]:
    """Build rows for a session (begin group, fields, end group)."""
    rows: list[dict] = []

    name = session.get("name", "session")
    label = session.get("label")
    condition = session.get("condition")

    # begin group
    begin = {"type": "begin group", "name": name, "appearance": "field-list"}
    begin.update(label_for_langs(label, languages))
    if condition:
        begin["relevant"] = condition
    rows.append(begin)

    # fields
    for field in session.get("fields", []):
        rows.append(build_field_row(field, languages))

    # end group
    rows.append({"type": "end group", "name": name})

    return rows


def build_group_rows(group: dict, languages: list[str]) -> list[dict]:
    """Build rows for a design group (begin group, sessions, end group)."""
    rows: list[dict] = []

    name = group.get("name", "group")
    label = group.get("label")

    # begin group
    begin = {"type": "begin group", "name": name}
    begin.update(label_for_langs(label, languages))
    rows.append(begin)

    # sessions inside the group
    for session in group.get("sessions", []):
        rows.extend(build_session_rows(session, languages))

    # end group
    rows.append({"type": "end group", "name": name})

    return rows


def build_summary_rows(summary_notes: list[dict], languages: list[str]) -> list[dict]:
    """Build the summary group at the end of the form."""
    rows: list[dict] = []

    # begin group summary
    begin = {"type": "begin group", "name": "summary",
             "appearance": "field-list summary"}
    begin.update(no_label(languages))
    rows.append(begin)

    for note in summary_notes:
        row = {"type": "note", "name": note.get("name", "note")}
        note_label = note.get("label") or note.get("labels")
        row.update(label_for_langs(note_label, languages))
        if note.get("relevant"):
            row["relevant"] = note["relevant"]
        rows.append(row)

    # end group summary
    rows.append({"type": "end group", "name": "summary"})

    return rows


def build_all_survey_rows(design: dict, languages: list[str]) -> list[dict]:
    """Assemble the full survey row list in the correct order."""
    rows: list[dict] = []

    # 1. CHT inputs group
    rows.extend(build_inputs_group(languages))

    # 2. Standard calculate fields
    rows.extend(build_standard_calculates())

    # 3. Contact-summary calculate fields
    cs_vars = design.get("contact_summary_calculates", [])
    if cs_vars:
        rows.extend(build_contact_summary_calculates(cs_vars))

    # 4. Inferred calculate fields
    inferred = design.get("inferred_calculates", [])
    if inferred:
        rows.extend(build_inferred_calculates(inferred))

    # 5. Design groups -> sessions -> fields
    for group in design.get("groups", []):
        rows.extend(build_group_rows(group, languages))

    # 6. Summary notes (if provided)
    summary_notes = design.get("summary_notes")
    if summary_notes:
        rows.extend(build_summary_rows(summary_notes, languages))

    return rows


# ---------------------------------------------------------------------------
# Choices sheet
# ---------------------------------------------------------------------------

def build_all_choices_rows(design: dict, languages: list[str]) -> list[dict]:
    """Build all choices rows from the design's choice_lists."""
    rows: list[dict] = []

    choice_lists = design.get("choice_lists", {})

    for list_name, options in choice_lists.items():
        for option in options:
            row = {"list_name": list_name}

            if isinstance(option, dict):
                row["name"] = option.get("name", option.get("value", ""))
                opt_label = option.get("label")
                row.update(label_for_langs(opt_label, languages))
            else:
                # Simple string option
                row["name"] = str(option)
                for lang in languages:
                    row[f"label::{lang}"] = str(option)

            rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Settings sheet
# ---------------------------------------------------------------------------

def build_settings_row(design: dict, form_id: str | None, form_title: str | None) -> dict:
    """Build the single settings row."""
    # form_title: CLI override > design.form_title (dict or string) > fallback
    if form_title:
        title = form_title
    else:
        raw_title = design.get("form_title", design.get("title", "Form"))
        if isinstance(raw_title, dict):
            # Use first available language value
            title = next(iter(raw_title.values()), "Form")
        else:
            title = str(raw_title)

    # form_id: CLI override > design.form_id > fallback
    fid = form_id or design.get("form_id", design.get("id", "form"))

    return {
        "form_title": title,
        "form_id": fid,
        "version": "now()",
        "style": "pages",
    }


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_sheet(ws, columns: list[str], rows: list[dict]):
    """Write header + data rows to a worksheet with auto-width."""
    # Header row (bold)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            if value:  # Only write non-empty
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width (capped at 50)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


def generate_xlsform(design: dict, output_path: str, languages: list[str],
                     form_id: str | None = None, form_title: str | None = None):
    """Generate the complete XLSForm workbook."""
    wb = Workbook()

    # --- survey sheet ---
    ws_survey = wb.active
    ws_survey.title = "survey"
    s_cols = survey_columns(languages)
    s_rows = build_all_survey_rows(design, languages)
    write_sheet(ws_survey, s_cols, s_rows)

    # --- choices sheet ---
    ws_choices = wb.create_sheet("choices")
    c_cols = choices_columns(languages)
    c_rows = build_all_choices_rows(design, languages)
    write_sheet(ws_choices, c_cols, c_rows)

    # --- settings sheet ---
    ws_settings = wb.create_sheet("settings")
    st_cols = settings_columns()
    st_row = build_settings_row(design, form_id, form_title)
    write_sheet(ws_settings, st_cols, [st_row])

    # Save
    wb.save(output_path)
    print(f"Generated XLSForm: {output_path}", file=sys.stderr)
    print(f"  Survey rows:  {len(s_rows)}", file=sys.stderr)
    print(f"  Choice rows:  {len(c_rows)}", file=sys.stderr)
    print(f"  Languages:    {', '.join(languages)}", file=sys.stderr)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate CHT XLSForm Excel files from processed design JSON",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    uv run scripts/generate-xlsform.py input.json output.xlsx
    uv run scripts/generate-xlsform.py input.json output.xlsx --languages fr,en
    uv run scripts/generate-xlsform.py input.json output.xlsx --form-id patient_assessment_under_5
        """,
    )
    parser.add_argument("input_json", help="Path to the input JSON file (from parse-design-sheet.py)")
    parser.add_argument("output_xlsx", help="Path to the output XLSForm Excel file")
    parser.add_argument("--languages", default=None,
                        help="Comma-separated languages (default: from JSON or fr,en)")
    parser.add_argument("--form-id", default=None, help="Override form_id in settings")
    parser.add_argument("--form-title", default=None, help="Override form_title in settings")
    return parser.parse_args()


def load_json(path: str) -> dict:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Error: file not found: {path}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: invalid JSON in {path}: {e}", file=sys.stderr)
        sys.exit(1)


def main():
    args = parse_args()

    design = load_json(args.input_json)

    # Resolve languages: CLI flag > JSON field > default
    if args.languages:
        languages = [lang.strip() for lang in args.languages.split(",")]
    elif "languages" in design:
        languages = design["languages"]
    else:
        languages = ["fr", "en"]

    generate_xlsform(design, args.output_xlsx, languages,
                     form_id=args.form_id, form_title=args.form_title)


if __name__ == "__main__":
    main()
